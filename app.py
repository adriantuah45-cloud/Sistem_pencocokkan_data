"""
=========================================================================
 SISTEM: Import File A & File B -> Isi Unit Kerja Otomatis
=========================================================================
Cara menjalankan (di VS Code / terminal):

    pip install streamlit pandas openpyxl
    streamlit run app.py

Setelah itu browser akan otomatis terbuka (biasanya di
http://localhost:8501). Di situ kamu tinggal:
  1. Upload File A (yang belum ada Unit Kerja)
  2. Upload File B (yang sudah ada Unit Kerja)
  3. Pilih kolom NIP di masing-masing file & kolom Unit Kerja di File B
  4. Klik "Proses"
  5. Download hasilnya (Hasil_File_A.xlsx)

Tidak perlu edit kode sama sekali.
=========================================================================
"""

import io
import re
import pandas as pd
import streamlit as st

st.set_page_config(page_title="Isi Unit Kerja Otomatis", layout="wide")

st.title("📋 Sistem Isi Unit Kerja Otomatis (berdasarkan NIP)")
st.caption(
    "Upload File A (tanpa Unit Kerja) dan File B (dengan Unit Kerja), "
    "sistem akan mencocokkan otomatis berdasarkan NIP."
)

# ---------------------------------------------------------------------------
def bersihkan_nip(series):
    """Bikin NIP jadi string angka murni (hilangkan tanda kutip ', spasi, strip,
    .0 dari angka float, dan karakter non-digit lainnya) supaya pencocokan akurat
    walau format asal beda-beda antar file."""
    def clean(x):
        if pd.isna(x):
            return ""
        s = str(x).strip()
        s = re.sub(r"\.0$", "", s)     # buang .0 kalau NIP kebaca sebagai angka desimal
        s = re.sub(r"[^0-9]", "", s)   # sisakan angka saja: buang ' spasi - dsb
        return s
    return series.apply(clean)


def pilih_engine(nama_file):
    """Pilih engine baca sesuai ekstensi file: .xls pakai xlrd, .xlsx pakai openpyxl."""
    if nama_file.lower().endswith(".xls"):
        return "xlrd"
    return "openpyxl"


def is_csv(nama_file):
    return nama_file.lower().endswith(".csv")


def is_html_menyamar_xls(file_bytes):
    """Banyak sistem (mis. laporan DUK BKN) 'export ke Excel' tapi isinya HTML biasa
    yang cuma diberi nama .xls. Deteksi dari beberapa byte pertama file."""
    awal = file_bytes[:200].lstrip().lower()
    return awal.startswith(b"<html") or awal.startswith(b"<!doctype") or b"<table" in awal[:500]


@st.cache_data
def intip_sheet_names(file_bytes, nama_file):
    if is_csv(nama_file):
        return ["(file csv)"]
    if is_html_menyamar_xls(file_bytes):
        return ["(tabel HTML)"]
    engine = pilih_engine(nama_file)
    xls = pd.ExcelFile(io.BytesIO(file_bytes), engine=engine)
    return xls.sheet_names


def baca_csv_fleksibel(file_bytes, header_row=0):
    """Coba beberapa delimiter & encoding umum di Indonesia (koma, titik koma, tab)."""
    for encoding in ["utf-8-sig", "utf-8", "latin1", "cp1252"]:
        for sep in [",", ";", "\t"]:
            try:
                df = pd.read_csv(
                    io.BytesIO(file_bytes),
                    header=header_row,
                    dtype=str,
                    sep=sep,
                    encoding=encoding,
                )
                if df.shape[1] > 1:  # kalau cuma 1 kolom, kemungkinan sep salah
                    return df
            except Exception:
                continue
    # fallback terakhir kalau semua percobaan di atas gagal
    return pd.read_csv(io.BytesIO(file_bytes), header=header_row, dtype=str)


@st.cache_data
def baca_excel(file_bytes, nama_file, sheet_name=0, header_row=0):
    if is_csv(nama_file):
        return baca_csv_fleksibel(file_bytes, header_row=header_row)

    if is_html_menyamar_xls(file_bytes):
        # File sebenarnya HTML (hasil export sistem web), bukan file Excel asli
        tabel_list = pd.read_html(io.BytesIO(file_bytes), header=header_row)
        # Ambil tabel terbesar (paling banyak barisnya) -> biasanya itu tabel datanya
        df = max(tabel_list, key=lambda t: t.shape[0])
        return df.astype(str)

    engine = pilih_engine(nama_file)
    return pd.read_excel(
        io.BytesIO(file_bytes),
        sheet_name=sheet_name,
        header=header_row,
        dtype=str,
        engine=engine,
    )


def ke_excel_bytes(df):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Hasil")
        ws = writer.sheets["Hasil"]

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell_font = Font(name="Calibri", size=11)
        thin = Side(style="thin", color="B7B7B7")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Format header
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Format isi + hitung lebar kolom otomatis
        for col_idx, col_name in enumerate(df.columns, start=1):
            panjang_maks = len(str(col_name))
            for row_idx in range(len(df)):
                nilai = df.iloc[row_idx, col_idx - 1]
                nilai_str = "" if pd.isna(nilai) else str(nilai)
                panjang_maks = max(panjang_maks, len(nilai_str))

                cell = ws.cell(row=row_idx + 2, column=col_idx)
                cell.font = cell_font
                cell.border = border
                cell.alignment = Alignment(vertical="center")

            lebar = min(max(panjang_maks + 3, 10), 45)
            ws.column_dimensions[get_column_letter(col_idx)].width = lebar

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28

    return buf.getvalue()


# ---------------------------------------------------------------------------
col1, col2 = st.columns(2)

with col1:
    st.subheader("1️⃣ File A (belum ada Unit Kerja)")
    file_a = st.file_uploader("Upload File A (.xlsx, .xls, atau .csv)", type=["xlsx", "xls", "csv"], key="file_a")

with col2:
    st.subheader("2️⃣ File B (sudah ada Unit Kerja)")
    file_b = st.file_uploader("Upload File B (.xlsx, .xls, atau .csv)", type=["xlsx", "xls", "csv"], key="file_b")

if file_a and file_b:
    bytes_a = file_a.getvalue()
    bytes_b = file_b.getvalue()

    st.divider()
    st.subheader("2.5️⃣ Atur sheet & baris header (kalau file punya judul laporan di atas tabel)")
    c_sheet_a, c_header_a, c_sheet_b, c_header_b = st.columns(4)

    with c_sheet_a:
        sheets_a = intip_sheet_names(bytes_a, file_a.name)
        sheet_a = st.selectbox("Sheet File A", sheets_a, index=0, key="sheet_a")
    with c_header_a:
        header_a = st.number_input(
            "Baris header File A (1 = baris pertama)", min_value=1, value=1, step=1, key="header_a"
        )
    with c_sheet_b:
        sheets_b = intip_sheet_names(bytes_b, file_b.name)
        sheet_b = st.selectbox("Sheet File B", sheets_b, index=0, key="sheet_b")
    with c_header_b:
        header_b = st.number_input(
            "Baris header File B (1 = baris pertama)", min_value=1, value=1, step=1, key="header_b"
        )

    st.caption(
        "Lihat contoh File B kamu: kalau header tabel (NO, NAMA, NIP, dst) ada di baris ke-11 "
        "Excel, isi '11' di kolom Baris header File B."
    )

    df_a = baca_excel(bytes_a, file_a.name, sheet_name=sheet_a, header_row=header_a - 1)
    df_b = baca_excel(bytes_b, file_b.name, sheet_name=sheet_b, header_row=header_b - 1)

    # Buang kolom "Unnamed: x" kosong hasil merge cell yang tidak terbaca rapi
    df_a = df_a.loc[:, ~df_a.columns.astype(str).str.startswith("Unnamed")]
    df_b = df_b.loc[:, ~df_b.columns.astype(str).str.startswith("Unnamed")]

    st.write("Pratinjau File A:")
    st.dataframe(df_a.head(5), use_container_width=True)
    st.write("Pratinjau File B:")
    st.dataframe(df_b.head(5), use_container_width=True)

    st.divider()
    st.subheader("3️⃣ Pilih kolom yang sesuai")

    c1, c2, c3 = st.columns(3)
    with c1:
        kolom_nip_a = st.selectbox("Kolom NIP di File A", df_a.columns, key="nip_a")
    with c2:
        kolom_nip_b = st.selectbox("Kolom NIP di File B", df_b.columns, key="nip_b")
    with c3:
        default_uk = 0
        for i, c in enumerate(df_b.columns):
            if "unit" in c.lower():
                default_uk = i
                break
        kolom_uk_b = st.selectbox(
            "Kolom Unit Kerja di File B", df_b.columns, index=default_uk, key="uk_b"
        )

    nama_kolom_baru = st.text_input("Nama kolom Unit Kerja di hasil File A", value="UNIT_KERJA")

    st.divider()

    if st.button("🚀 Proses", type="primary"):
        df_a_work = df_a.copy()
        df_b_work = df_b.copy()

        df_a_work["_NIP_CLEAN"] = bersihkan_nip(df_a_work[kolom_nip_a])
        df_b_work["_NIP_CLEAN"] = bersihkan_nip(df_b_work[kolom_nip_b])

        df_b_unik = df_b_work.drop_duplicates(subset="_NIP_CLEAN", keep="first")
        mapping = dict(zip(df_b_unik["_NIP_CLEAN"], df_b_unik[kolom_uk_b]))

        df_a_work[nama_kolom_baru] = df_a_work["_NIP_CLEAN"].map(mapping)

        tidak_cocok = df_a_work[df_a_work[nama_kolom_baru].isna()].copy()
        tidak_cocok = tidak_cocok.drop(columns=["_NIP_CLEAN", nama_kolom_baru])

        hasil = df_a_work.drop(columns=["_NIP_CLEAN"])

        total = len(hasil)
        cocok = total - len(tidak_cocok)

        st.success("Selesai diproses!")
        m1, m2, m3 = st.columns(3)
        m1.metric("Total baris File A", total)
        m2.metric("Berhasil dicocokkan", cocok)
        m3.metric("Tidak ditemukan", len(tidak_cocok))

        st.subheader("📄 Pratinjau Hasil")
        st.dataframe(hasil, use_container_width=True)

        st.download_button(
            "⬇️ Download Hasil_File_A.xlsx",
            data=ke_excel_bytes(hasil),
            file_name="Hasil_File_A.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        if len(tidak_cocok) > 0:
            st.subheader("⚠️ NIP yang Tidak Ditemukan di File B")
            st.dataframe(tidak_cocok, use_container_width=True)
            st.download_button(
                "⬇️ Download Laporan_Tidak_Cocok.xlsx",
                data=ke_excel_bytes(tidak_cocok),
                file_name="Laporan_Tidak_Cocok.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
else:
    st.info("⬆️ Silakan upload File A dan File B terlebih dahulu untuk mulai.")